"""
Unified Excel reader.

The "xlsx" files published by some issuers (observed e.g. with Amundi
files) contain slightly invalid internal XML (colors not in aRGB format)
that makes ``openpyxl`` fail. We therefore use ``python-calamine`` (a Rust
parser, very tolerant, supports xls/xlsx/xlsb/ods) as the primary engine,
with a fallback to a dedicated parser for the "SpreadsheetML" format
(Excel 2003 XML, typical of some BlackRock/iShares files with a .xls
extension) and finally to ``openpyxl`` for the rare cases where calamine
is not enough.

The output is always normalized to: Dict[str(sheet_name), List[List[Any]]]
"""
from __future__ import annotations
from typing import Dict, List, Union
from pathlib import Path
import io

from ..exceptions import UnreadableFileError
from . import spreadsheetml


def _read_head(path_or_bytes: Union[str, bytes]) -> bytes:
    if isinstance(path_or_bytes, (bytes, bytearray)):
        return bytes(path_or_bytes[:4096])
    with open(path_or_bytes, "rb") as f:
        return f.read(4096)


def _read_with_calamine(path_or_bytes) -> Dict[str, List[list]]:
    from python_calamine import CalamineWorkbook

    if isinstance(path_or_bytes, (bytes, bytearray)):
        wb = CalamineWorkbook.from_filelike(io.BytesIO(path_or_bytes))
    else:
        wb = CalamineWorkbook.from_path(str(path_or_bytes))

    sheets = {}
    for name in wb.sheet_names:
        sheets[name] = wb.get_sheet_by_name(name).to_python()
    return sheets


def _read_with_openpyxl(path_or_bytes) -> Dict[str, List[list]]:
    import openpyxl

    target = io.BytesIO(path_or_bytes) if isinstance(path_or_bytes, (bytes, bytearray)) else path_or_bytes
    wb = openpyxl.load_workbook(target, data_only=True, read_only=True)
    sheets = {}
    for name in wb.sheetnames:
        ws = wb[name]
        sheets[name] = [list(row) for row in ws.iter_rows(values_only=True)]
    return sheets


def _read_with_spreadsheetml(path_or_bytes) -> Dict[str, List[list]]:
    if isinstance(path_or_bytes, (bytes, bytearray)):
        import xml.etree.ElementTree as ET
        # spreadsheetml.read_workbook expects a path; parse via a temp file
        # so we can reuse the module's implementation as-is.
        tree = ET.fromstring(path_or_bytes)
        import tempfile
        with tempfile.NamedTemporaryFile(suffix=".xml") as tmp:
            tmp.write(path_or_bytes)
            tmp.flush()
            return spreadsheetml.read_workbook(tmp.name)
    return spreadsheetml.read_workbook(str(path_or_bytes))


def read_workbook(path_or_bytes: Union[str, "Path", bytes]) -> Dict[str, List[list]]:
    """
    Reads an Excel file (xls/xlsx/xlsb/ods, or a pseudo-.xls SpreadsheetML
    XML file) and returns {sheet_name: row_matrix}.

    Tries, in order: calamine -> spreadsheetml (if the sniff detects XML)
    -> openpyxl. Raises UnreadableFileError if all of them fail.
    """
    if isinstance(path_or_bytes, Path):
        path_or_bytes = str(path_or_bytes)

    head = _read_head(path_or_bytes)
    errors = []

    if spreadsheetml.sniff(head):
        try:
            return _read_with_spreadsheetml(path_or_bytes)
        except Exception as e:  # pragma: no cover - fallback path
            errors.append(f"spreadsheetml: {e!r}")

    try:
        return _read_with_calamine(path_or_bytes)
    except Exception as e:
        errors.append(f"calamine: {e!r}")

    try:
        return _read_with_openpyxl(path_or_bytes)
    except Exception as e:
        errors.append(f"openpyxl: {e!r}")

    try:
        return _read_with_spreadsheetml(path_or_bytes)
    except Exception as e:
        errors.append(f"spreadsheetml: {e!r}")

    raise UnreadableFileError(
        "Could not read the file with any of the available readers:\n"
        + "\n".join(errors)
    )
